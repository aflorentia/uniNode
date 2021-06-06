#Basic Libraries
from tkinter import *
import tkinter as tk
from PIL import Image, ImageTk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter.font import Font
from tkinter import ttk,filedialog
from statistics import mean
from tkinter import messagebox


#Dispay setting 
screen = Tk()
screen.minsize(600,600)
screen.title("UniNode")
screen.iconbitmap('logo.ico')
screen.configure(background = "white")
my_img=ImageTk.PhotoImage(Image.open("logo.png"))
my_label=Label(image=my_img)
my_label.pack()
custFont = Font(
    family = "Helvetica",
    size = 24,
    weight = "bold"
)
title= Label (text="Calculate Grades", font = custFont,bg="white", fg =  "#800000")
title.place(x=170, y=80)   

#Select Courses     
wb = Workbook()
wb =load_workbook('averageGrade.xlsx')
ws = wb.active
column_a = ws['A']
column_b = ws ['B']

SelectCourse =  Label(screen,text="Select Course",bg="white",fg ="#800000")
SelectCourse.place(x=380,y=180)

my_frame = Frame(screen)
my_scrollbar = Scrollbar(my_frame,orient=VERTICAL)

 
my_listbox = Listbox(my_frame,width=45,yscrollcommand = my_scrollbar.set)

my_scrollbar.config(command=my_listbox.yview)
my_scrollbar.pack(side=RIGHT,fill=Y)
my_frame.place(x=290,y=200)

my_listbox.pack()

my_list = column_a

for item in my_list:
    my_listbox.insert(END,item.value)
    
global selected     

selected = Label(screen,text='',bg="white") 
selected.place(x=300, y =390)    
    
    
def select(): 
    selected.config(text=my_listbox.get(ANCHOR))
    my_listbox.delete(ANCHOR)

SelectButton = Button(screen,text = "OK!",fg="white",bg ="#800000",command=select)
SelectButton.place(x=550,y=365)
 

#To pick grade
AddGrade =  Label(screen,text="Add Grade",bg="white",fg ="#800000")
AddGrade.place(x=380,y=440)    
options = [
    5,
    6,
    7,
    8,
    9,
    10
]
clicked = StringVar()
clicked.set(options[0])
drop = OptionMenu(screen,clicked,*options)
drop.place(x=380,y=470)

thislist = []


#To instert course and grade to Courses-Grade list
Courses_Grade = Label(screen,text = "Courses      " + '\t\t\t' + "      Grade",bg="#800000",fg="white")
Courses_Grade.place(x=20,y=200) 

frame = Frame(screen)
scrollbar = Scrollbar(frame,orient=HORIZONTAL)

listbox1 = Listbox(frame,width=41,xscrollcommand = scrollbar.set)


scrollbar.config(command=listbox1.xview)
scrollbar.pack(side=BOTTOM,fill=X)
frame.place(x=18,y=221)

listbox1.pack()


#Exception 
def error():
    messagebox.showerror("Warning Message","You need to select a course")
    



def Insert():
    course_grade = selected.cget("text")+"                   " + clicked.get()
    if len(selected.cget("text")) == 0:
        error()
    else:
        listbox1.insert(END,course_grade)
        thislist.append(clicked.get())
        mylist = list(map(int,thislist))
        average = mean(mylist)
        avg.config(text=average)
        selected.config(text='')
        
     
    
    
Insert = Button(screen,text="Add",bg="#800000",fg="white",command=Insert)
Insert.place(x=230,y=450)    



avg = Label(screen,text='',bg="#800000",fg="white")
avg.place(x=70, y =400)

Courses_Grade = Label(screen,text = "Average",bg="#800000",fg="white")
Courses_Grade.place(x=20,y=400)  


#Exit Button
Quit= Button(screen ,text= "Exit Program",width = 18 ,height = 2 ,fg ="#800000", command = screen.quit)
Quit.place(x= 220, y=550)


screen.mainloop() 
