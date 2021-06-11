import numpy as np
import openpyxl
import pandas as pd
from tkinter import *
from PIL import ImageTk, Image
from openpyxl.workbook import Workbook
from tkcalendar import *
from openpyxl import load_workbook
from tkinter.font import Font
from tkinter import ttk
from tkinter import messagebox
from tkinter.messagebox import showinfo

screen = Tk()
screen.minsize(600 , 600)
screen.maxsize(600 , 600)
screen.title("UniNode")
screen.configure(background = "white")

name_var =StringVar()
dat_var =StringVar()
tim_var =StringVar()

        
def showEvents(dat1):
    dat1 = dat1.replace('/', '.')
    screen1 = Tk()
    screen1.title("Event List")
    screen1.minsize(600,600)
    screen1.maxsize(600 , 600)


    wb = Workbook()
    wb =load_workbook('/Users/emmaeconomopoulou/Desktop/gui/events.xlsx')
    ws = wb.active

    cola= ws['A']
    colb= ws['B']
    colc= ws['C']
    label = Label(screen1, text="")
    label.place(x=200, y=200)
    list= ''

    df = pd.read_excel('/Users/emmaeconomopoulou/Desktop/gui/events.xls', index_col=None , header=None )
    for row in range(0, df.shape[0]):
        if dat1==df.loc[row][1]:
            msg = str(df.loc[row])
            label.config(text=msg)
        else:
            msg = "You are free to fly"
            label.config(text=msg)
    button_quitp= Button(screen1 ,text= "Ok",width = 18 ,height = 2 ,fg ="#800000", command = lambda :screen1.destroy())
    button_quitp.place(x= 220, y=500)


def showEvents(dat1):
    dat1 = dat1.replace('/', '.')
    screen1 = Tk()
    screen1.title("Event List")
    screen1.minsize(600,600)
    screen1.maxsize(600 , 600)


    wb = Workbook()
    wb =load_workbook('/Users/emmaeconomopoulou/Desktop/gui/events.xlsx')
    ws = wb.active

    cola= ws['A']
    colb= ws['B']
    colc= ws['C']
    label = Label(screen1, text="")
    label.pack(padx=150)
    list= ''

    df = pd.read_excel('/Users/emmaeconomopoulou/Desktop/gui/events.xls', index_col=None , header=None )
    for row in range(0, df.shape[0]):
        if dat1==df.loc[row][1]:
            msg = str(df.loc[row])
            label.config(text=msg)
        else:
            msg = "You are free to fly"
            label.config(text=msg)
    button_quitp= Button(screen1 ,text= "Ok",width = 18 ,height = 2 ,fg ="#800000", command = lambda :screen1.destroy())
    button_quitp.place(x= 220, y=500)


def submitData():
            name=name_var.get()
            dat=dat_var.get()
            tim=tim_var.get()

            path='/Users/emmaeconomopoulou/Desktop/gui/events.xlsx'
            wb =load_workbook(path)
            ws = wb.active
            ws.cell(column=1, row=ws.max_row+1, value=name )
            ws.cell(column=2, row=ws.max_row, value=dat )
            ws.cell(column=3, row=ws.max_row, value=tim )
            wb.save(path)

            messagebox.showinfo("Info", "Submitted")


def addEvent():
        
    screen1 = Toplevel(screen)
    screen1.title("Add Events")
    screen1.minsize(600,600)
    screen1.maxsize(600 , 600)

    custFont = Font(
            family = "Helvetica",
            size = 24,
            weight = "bold"
        )


    main_frame =Frame(screen1)
    main_frame.pack(fill=BOTH, expand=1)

    my_canvas= Canvas(main_frame)
    my_canvas.pack(side=LEFT, fill=BOTH , expand=1)

    my_scrollbar = ttk.Scrollbar(main_frame, orient= VERTICAL , command= my_canvas.yview)
    my_scrollbar.pack(side= RIGHT ,fill=Y)

    my_canvas.configure(yscrollcommand = my_scrollbar.set)
    my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion= my_canvas.bbox("all")))

    second_frame= Frame(my_canvas)

    my_canvas.create_window((250,300),window=second_frame )

    wb = Workbook()
    wb =load_workbook('/Users/emmaeconomopoulou/Desktop/gui/events.xlsx')
    ws = wb.active
    cola= ws['A']
    colb= ws['B']
    colc= ws['C']
    label = Label(second_frame, text="")
    label.pack(padx=150)
    list= ''
    for cell in ws.iter_rows(min_row=1, max_row=7,min_col=1,max_col=3,values_only=True ):
        list= f'{list +"●  " + str(cell) }\n'
        list = list.replace('(', '')
        list = list.replace(')', '')
        list = list.replace('\'', '')
        label.config(text=list)
        
    def add_event():
        screen2 = Toplevel(screen1)
        screen2.title("Add Events")
        screen2.minsize(600,600)
        screen2.maxsize(600 , 600)

        custFont = Font(
            family = "Helvetica",
            size = 24,
            weight = "bold"
        )

        title_e=Label(screen2, text= "Add a new Event" , fg="#800000", font= custFont)
        title_e.place(x=200, y =80)

        eventName=Entry(screen2,textvariable=name_var)
        eventName.place(x=200 , y=200)
        ex1=Label(screen2, text = "e.g. Data Bases Project", fg='grey')
        ex1.place(x=400, y=200)
        eventNameLab = Label(screen2, text = "Enter Event Name",fg="#800000")
        eventNameLab.place(x=235, y=170)

        eventDate=Entry(screen2,textvariable=dat_var)
        eventDate.place(x=200 , y=300)
        ex2=Label(screen2, text = "e.g. 5.8.22 ", fg='grey')
        ex2.place(x=400, y=300)
        eventDateLab = Label(screen2, text = "Enter Event Date",fg="#800000")
        eventDateLab.place(x=235, y=270)

        eventTime=Entry(screen2,textvariable=tim_var)
        eventTime.place(x=200 , y=400)
        ex3=Label(screen2, text = "e.g. 3.35 ", fg='grey')
        ex3.place(x=400, y=400)
        eventTimeLab = Label(screen2, text = "Enter Event Time", fg="#800000")
        eventTimeLab.place(x=235, y=370)

        button_submit= Button(screen2 ,text= "Submit",width = 8 ,height = 2 ,fg ="#800000", command = submitData)
        button_submit.place(x= 255, y=450)

        button_q= Button(screen2 ,text= "Back",width = 8 ,height = 2 ,fg ="#800000", command = lambda: screen2.destroy())
        button_q.place(x= 500, y=550)


    add_event_btn= Button(screen1 ,text= "Add Event",width = 8 ,height = 2 ,fg ="#800000", command = add_event)
    add_event_btn.place(x=255, y=150)

    title_e=Label(screen1, text= "Event Handling" , fg="#800000", font= custFont)
    title_e.place(x=200, y =80)

    button_quit= Button(screen1 ,text= "OK",width = 8 ,height = 2 ,fg ="#800000", command = lambda: screen1.destroy())
    button_quit.place(x= 500, y=550)


def selectCourses():

    screen1 = Toplevel(screen)
    screen1.title("Select Courses")
    screen1.minsize(600,600)

    main_frame =Frame(screen1)
    main_frame.pack(fill=BOTH, expand=1)

    my_canvas= Canvas(main_frame)
    my_canvas.pack(side=LEFT, fill=BOTH , expand=1)

    my_scrollbar = ttk.Scrollbar(main_frame, orient= VERTICAL , command= my_canvas.yview)
    my_scrollbar.pack(side= RIGHT ,fill=Y)

    my_canvas.configure(yscrollcommand = my_scrollbar.set)
    my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion= my_canvas.bbox("all")))

    second_frame= Frame(my_canvas)

    my_canvas.create_window((0,0),window=second_frame, anchor= "nw")

    wb = Workbook()
    wb =load_workbook('/Users/emmaeconomopoulou/Desktop/gui/courses.xlsx')
    ws = wb.active
    cola= ws['A']

    list1=''
    checkbox_vars = {}

    values=[0]*116
    def check():
        
        for i,cell in enumerate(cola):
            values[i]=(checkbox_vars[cell].get())
        print(values)
    for cell in cola:
        var=IntVar()
        list1= Checkbutton(second_frame, text= str(cell.value) ,variable=var, command = check)
        list1.pack()
        checkbox_vars[cell] = var
        
    # path1 = "/Users/emmaeconomopoulou/Desktop/gui/courses.xlsx"
    # wb1=openpyxl.load_workbook(path1)
    # ws1=wb1.active
    for row in ws[ 'b1' : 'b116']:
        for index ,cell in enumerate(row):
            cell.value = values[index]
    wb.save("/Users/emmaeconomopoulou/Desktop/gui/courses1.xlsx")
    button_quit1= Button(screen1 ,text= "OK",width = 8 ,height = 2 ,fg ="#800000", command =lambda: screen1.destroy())
    button_quit1.place(x= 500, y=550)
    
    

class CalendarMain():

    def __init__(self,master):
        myFrame = Frame(master)
        myFrame.pack()

        def selectDate():
            myDate=self.myCal.get_date()
            label_date= Label(screen,fg ="#800000", text="")
            label_date.place(x=155, y= 200)
            label_date.config(text=myDate)
            events= Button(screen, text= "Events",width = 5 ,height = 1 ,fg ="#800000", command =lambda :showEvents(myDate))
            events.place(x=220, y =202)


        self.myCal= Calendar (screen, setmode = 'day', date_pattern ='d/m/yy')
        self.myCal.place(x =70 , y = 250 )  

        custFont = Font(
            family = "Helvetica",
            size = 24,
            weight = "bold"
        )

        self.title= Label (text="Student Calendar", font = custFont, fg =  "#800000")
        self.title.place(x=197, y=130)

        self.openCal = Button(screen, text= "Select Courses",width = 18 ,height = 2 ,fg ="#800000", command = selectCourses)
        self.openCal.place(x=380, y =270)

        self.myDate= Button(screen, text= "Select Date",width = 8 ,height = 2 ,fg ="#800000", command =selectDate)
        self.myDate.place(x=140, y =430)

        self.label_date= Label(screen,fg ="#800000", text="")
        self.label_date.place(x=155, y= 200)

        self.openCal1 = Button(screen, text= "Add Event",width = 18 ,height = 2 ,fg ="#800000", command= addEvent)
        self.openCal1.place(x=380, y =360)

        self.button_quit= Button(screen ,text= "Exit Program",width = 18 ,height = 2 ,fg ="#800000", command= screen.quit)
        self.button_quit.place(x= 220, y=500)

        self.my_img=ImageTk.PhotoImage(Image.open("/Users/emmaeconomopoulou/Desktop/gui/logo.png"))
        self.my_label=Label(image=self.my_img)
        self.my_label.pack()
 
        
if __name__ == "__main__":
    c = CalendarMain(screen)

    screen.mainloop()