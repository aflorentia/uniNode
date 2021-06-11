from tkinter import *
import tkinter as tk
from PIL import Image, ImageTk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter.font import Font
from tkinter import ttk,filedialog
from statistics import mean
from tkinter import messagebox
class Calculate_Average_Grade:

    def main_screen(self):
        global screen
        global custFont
        global my_img
        screen = Tk()
        screen.geometry("600x600")
        screen.configure(background = "white")
        screen.title("UniNode")
        custFont =Font(
        family = "Helvetica",
        size = 24,
        weight = "bold"
        )
        screen.iconbitmap('logo.ico')
        my_img=ImageTk.PhotoImage(Image.open("logo.png"))
        my_label=Label(image=my_img,background="white")
        my_label.pack()
        title= Label(text="Calculate Grades", font = custFont,bg="white", fg =  "#800000")
        title.place(x=170, y=80) 
        #Select Courses     
        wb = Workbook()
        wb =load_workbook('averageGrade.xlsx')
        ws = wb.active
        column_a = ws['A']
        SelectCourse =  Label(text="Select Course",bg="white",fg ="#800000")
        SelectCourse.place(x=380,y=180)
        my_frame = Frame(screen)
        my_scrollbar = Scrollbar(my_frame,orient=VERTICAL)
        global my_listbox
        my_listbox = Listbox(my_frame,width=45,yscrollcommand = my_scrollbar.set)
        my_scrollbar.config(command=my_listbox.yview)
        my_scrollbar.pack(side=RIGHT,fill=Y)
        my_frame.place(x=290,y=200)
        my_listbox.pack()
        my_list = column_a
        for item in my_list:
            my_listbox.insert(END,item.value)
        SelectButton = Button(text = "OK!",fg="white",bg ="#800000",command=self.select)
        SelectButton.place(x=550,y=365)
        
        #Insert Grade
        AddGrade =  Label(text="Add Grade",bg="white",fg ="#800000")
        AddGrade.place(x=380,y=440)
        options = [
            5,
            6,
            7,
            8,
            9,
            10
        ]    
        global clicked
        clicked = StringVar()
        clicked.set(options[0])
        drop = OptionMenu(screen,clicked,*options)
        drop.place(x=380,y=470)
        global thislist 
        thislist = []
        
        #Courses-Grade list
        Courses_Grade = Label(screen,text = "Courses      " + '\t\t\t' + "      Grade",bg="#800000",fg="white")
        Courses_Grade.place(x=20,y=200)
        global listbox1
        listbox1 = Listbox(screen,width=41)
        listbox1.place(x=20,y=221)
        Insert = Button(text="Add",fg ="#800000",command=self.Insert)
        Insert.place(x=230,y=450)
        
        global avg
        avg = Label(text='',bg="#800000",fg="white")
        avg.place(x=70, y =385)
        Courses_Grade = Label(text = "Average",bg="#800000",fg="white")
        Courses_Grade.place(x=20,y=385)
        Quit= Button(screen ,text= "Exit Program",width = 18 ,height = 2 ,fg ="#800000", command = self.exit)
        Quit.place(x= 220, y=550)        
        
        screen.mainloop()
        
        
        
    def select(self):
        global selected
        selected = Label(text='',bg="white") 
        selected.place(x=300, y =390)
        selected.config(text=my_listbox.get(ANCHOR))
        my_listbox.delete(ANCHOR)        
        
        
        
    def error(self):
        messagebox.showerror("Warning Message","You need to select a course")
        
        
        
    def Insert(self):
        course_grade = selected.cget("text")+"                   " + clicked.get()
        if len(selected.cget("text")) == 0:
            error()
        else:
            listbox1.insert(END,course_grade)
            thislist.append(clicked.get())
            mylist = list(map(int,thislist))
            average = mean(mylist)
            avg.config(text=average)
            selected.config(text='')
            
            
    def exit(self):
        screen.destroy()
    
    
calc = Calculate_Average_Grade()
calc.main_screen()
